import openpyxl as px
import numpy as np
import random
from sklearn.ensemble import RandomForestClassifier

# read xlsx file
W = px.load_workbook('ANSUR.xlsx')
# read each sheet from excel file
pw = W.get_sheet_by_name(name = 'Women Analysis')
pm = W.get_sheet_by_name(name = 'Men Analysis')
# scrapping needed ratios
A=np.zeros([2208,8])
# mend data
ab=np.zeros([1774,5])
# women data
af=np.zeros([2208,5])
# women rating data
scoref=np.zeros([2208,5])
# men rating data
scoreb=np.zeros([1774,5])
# scrapping needed ratios for men  analysis data
rownum=0
colnum=0
for row in pw.iter_rows():
    if(rownum<3 or rownum>2210):
        rownum+=1
        continue
    for k in row:
        if(colnum==1):
            A[rownum - 3][0] = float(k.internal_value)
            colnum += 1
        elif(colnum==2):
            A[rownum - 3][1] = float(k.internal_value)
            colnum += 1
        elif (colnum == 5):
            A[rownum - 3][2] = float(k.internal_value)
            colnum += 1
        elif (colnum == 10):
            A[rownum - 3][3] = float(k.internal_value)
            colnum += 1
        elif (colnum == 18):
            A[rownum - 3][4] = float(k.internal_value)
            colnum += 1
        elif (colnum == 19):
            A[rownum - 3][5] = float(k.internal_value)
            colnum += 1
        elif (colnum == 21):
            A[rownum - 3][6] = float(k.internal_value)
            colnum += 1
        elif (colnum == 9):
            A[rownum - 3][7] = float(k.internal_value)
            colnum += 1
        else:
            colnum+=1
            continue
    val=A[rownum - 3]
    af[rownum - 3][0]=val[1]/val[0]#Leg/Body
    scoref[rownum - 3][0]=int(random.randint(0,21))
    af[rownum - 3][1] = val[2] / val[3]#Shoulder/Hip
    scoref[rownum - 3][1] = int(random.randint(0, 21))
    af[rownum - 3][2] = val[4] / val[5]#Chest/Waist
    scoref[rownum - 3][2] = int(random.randint(0, 21))
    af[rownum - 3][3] = val[5] / val[6]#Waist/Hip
    scoref[rownum - 3][3] = int(random.randint(0, 21))
    af[rownum - 3][4] = val[0] / val[7]#Body/Waist
    scoref[rownum - 3][4] = int(random.randint(0, 21))
    colnum = 0
    rownum+=1
# scrapping needed ratios for women  analysis data
rownum=0
colnum=0
for row in pm.iter_rows():
    if(rownum<1 or rownum>1774):
        rownum+=1
        continue
    for k in row:
        if (colnum == 1):
            val1 = float(k.internal_value)
            colnum += 1
        if(colnum==4):
            ab[rownum - 1][0] = float(k.internal_value)
            scoreb[rownum - 1][0] = int(random.randint(0, 21))
            colnum += 1
        elif(colnum==7):
            ab[rownum - 1][1] = float(k.internal_value)
            scoreb[rownum - 1][1] =int( random.randint(0, 21))
            colnum += 1
        elif (colnum == 10):
            ab[rownum - 1][2] = val1/float(k.internal_value)
            scoreb[rownum - 1][2] =int(random.randint(0, 21))
            colnum += 1
        elif (colnum == 23):
            ab[rownum - 1][3] = float(k.internal_value)
            scoreb[rownum - 1][3] = int(random.randint(0, 21))
            colnum += 1
        elif (colnum == 21):
            ab[rownum - 1][4] = float(k.internal_value)
            scoreb[rownum - 1][4] = int(random.randint(0, 21))
            colnum += 1
        else:
            colnum+=1
            continue
    colnum = 0
    rownum+=1

# randorm forest algorithm
def random_forest_classifier(features, target):
    clf = RandomForestClassifier()
    clf.fit(features, target)
    return clf

# sex=1 for male, sex=0 for female
sex=0
test=[0.5239,2.0802,1.1129,6.134,0.8408]# testing data
prediction=[]
if(sex==0):
    for i in range(5):
        train_x=af[:,i]# training data
        train_x=train_x.reshape(-1,1)
        train_y=scoref[:,i]# label data
        trained_model = random_forest_classifier(train_x, train_y)#  creat random forest model
        predictions = trained_model.predict(test[i])# predict rating for teat data
        prediction.append(predictions[0]*0.5)
else:
    for i in range(5):
        train_x=ab[:,i]# training data
        train_x = train_x.reshape(-1, 1)# label data
        train_y=scoreb[:,i]
        trained_model = random_forest_classifier(train_x, train_y)#  creat random forest model
        predictions = trained_model.predict(test[i])# predict rating for teat data
        prediction.append(predictions[0]*0.5)
# display results
M=np.mean(prediction)
print(prediction)
print(M)
