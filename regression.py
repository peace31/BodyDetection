import openpyxl as px
import numpy as np

# read  real xlsx file
Wb = px.load_workbook('REALDATA_boy.xlsx')
Wg = px.load_workbook('REALDATA - girl.xlsx')
# read each sheet from excel file
pb = Wb.get_sheet_by_name(name = 'Sheet1')
pf = Wg.get_sheet_by_name(name = 'Sheet1')


# read  measured xlsx file(result from code)
Wbb=px.load_workbook('MeasureLDATA_boy_back.xlsx')
Wfb=px.load_workbook('MeasureLDATA_boy.xlsx')
pbb = Wbb.get_sheet_by_name(name = 'Sheet1')# back  boy data
pfb = Wfb.get_sheet_by_name(name = 'Sheet1')# front boy data
# women data
Wbg=px.load_workbook('MeasureDATA - girl - back.xlsx')
Wfg=px.load_workbook('MeasureDATA - girl.xlsx')
pbg = Wbg.get_sheet_by_name(name = 'Sheet1')# back girls data
pfg = Wfg.get_sheet_by_name(name = 'Sheet1')# front boy data
# scrapping needed ratios
A=np.zeros([21,9])
ab=np.zeros([21,7])
af=np.zeros([50,8])
rownum=0
colnum=0
for row in pb.iter_rows():
    if(rownum==0):
        rownum+=1
        continue
    for k in row:
        if(colnum==0 or colnum>9):
            colnum+=1
            continue
        A[rownum - 1][colnum - 1] =float(k.internal_value)
        colnum+=1
    val=A[rownum - 1]
    ab[rownum - 1][0]=val[6]/val[0]
    ab[rownum - 1][1] = val[1] / val[7]
    ab[rownum - 1][2] = val[1] / val[8]
    ab[rownum - 1][3] = val[3] / val[2]
    ab[rownum - 1][4] = val[0] / val[2]*2.0
    ab[rownum - 1][5] = val[2] / val[7]/2.0
    ab[rownum - 1][6] = val[2] / val[8]/2.0
    colnum = 0
    rownum+=1
rownum=0
colnum=0
A=np.zeros([50,10])
for row in pf.iter_rows():
    if(rownum==0):
        rownum+=1
        continue
    for k in row:
        if(colnum==0 or colnum>10):
            colnum+=1
            continue
        A[rownum - 1][colnum - 1] =float(k.internal_value)
        colnum+=1
    val=A[rownum - 1]
    af[rownum - 1][0]=val[7]/val[0]
    af[rownum - 1][1] = val[5] / val[8]
    af[rownum - 1][2] = val[5] / val[9]
    af[rownum - 1][3] = val[3] / val[2]*2.0
    af[rownum - 1][4] = val[4] / val[2] * 2.0
    af[rownum - 1][5] = val[0] / val[2]*2.0
    af[rownum - 1][6] = val[2] / val[8]/2.0
    af[rownum - 1][7] = val[2] / val[9]/2.0
    colnum = 0
    rownum+=1


abb=np.zeros([21,5])
afb=np.zeros([21,5])

abg=np.zeros([50,5])
afg=np.zeros([50,5])
rownum=0
colnum=0
for row in pbb.iter_rows():
    if(rownum==0):
        rownum+=1
        continue
    for k in row:
        if(colnum==0 or colnum>5):
            colnum+=1
            continue
        abb[rownum - 1][colnum - 1] =float(k.internal_value)
        colnum+=1
    rownum+=1
    colnum = 0
rownum=0
colnum=0
for row in pfb.iter_rows():
    if(rownum==0):
        rownum+=1
        continue
    for k in row:
        if(colnum==0 or colnum>5):
            colnum+=1
            continue
        afb[rownum - 1][colnum - 1] =float(k.internal_value)
        colnum+=1
    rownum+=1
    colnum = 0
rownum=0
colnum=0
for row in pbg.iter_rows():
    if(rownum==0):
        rownum+=1
        continue
    for k in row:
        if(colnum==0 or colnum>5):
            colnum+=1
            continue
        abg[rownum - 1][colnum - 1] =float(k.internal_value)
        colnum+=1
    rownum+=1
    colnum = 0
rownum=0
colnum=0
for row in pfg.iter_rows():
    if(rownum==0):
        rownum+=1
        continue
    for k in row:
        if(colnum==0 or colnum>5):
            colnum+=1
            continue
        afg[rownum - 1][colnum - 1] =float(k.internal_value)
        colnum+=1
    rownum+=1
    colnum = 0
# regression method
lamdabb=np.zeros(5)
lamdafb=np.zeros(5)
M1=np.mean(ab[:,0])
M2=np.mean(ab[:,1])
M3=np.mean(ab[:,2])
M4=np.mean(ab[:,3])
M5=np.mean(ab[:,4])
M6=np.mean(ab[:,5])
M7=np.mean(ab[:,6])
M11=np.mean(abb[:,0])
M21=np.mean(abb[:,1])
M31=np.mean(abb[:,2])
M41=np.mean(abb[:,3])
M51=np.mean(abb[:,4])
M111=np.mean(afb[:,0])
M211=np.mean(afb[:,1])
M311=np.mean(afb[:,2])
M411=np.mean(afb[:,3])
M511=np.mean(afb[:,4])
S1=0;S2=0;S3=0;S4=0;S5=0
S11=0;S21=0;S31=0;S41=0;S51=0
T1=0;T2=0;T3=0;T4=0;T5=0
T11=0;T21=0;T31=0;T41=0;T51=0
for i in range(21):
    S1+=(ab[i][0]-M1)*(abb[i][0]-M11)
    S11+=(abb[i][0]-M11)*(abb[i][0]-M11)
    S2 += (ab[i][1] - M2) * (abb[i][1] - M21)
    S21 += (abb[i][1] - M21) * (abb[i][1] - M21)
    S3 += (ab[i][3] - M4) * (abb[i][2] - M31)
    S31 += (abb[i][2] - M31) * (abb[i][2] - M31)
    S4 += (ab[i][4] - M5) * (abb[i][3] - M41)
    S41 += (abb[i][3] - M41) * (abb[i][3] - M41)
    S5 += (ab[i][5] - M6) * (abb[i][4] - M51)
    S51 += (abb[i][4] - M51) * (abb[i][4] - M51)

    T1 += (ab[i][0] - M1) * (afb[i][0] - M111)
    T11 += (afb[i][0] - M111) * (afb[i][0] - M111)
    T2 += (ab[i][2] - M3) * (afb[i][1] - M211)
    T21 +=  (afb[i][1] - M211) *  (afb[i][1] - M211)
    T3 += (ab[i][3] - M4) * (afb[i][2] - M311)
    T31 += (afb[i][2] - M311) * (afb[i][2] - M311)
    T4 += (ab[i][4] - M5) * (afb[i][3] - M411)
    T41 += (afb[i][3] - M411) * (afb[i][3] - M411)
    T5 += (ab[i][6] - M7) * (afb[i][4] - M511)
    T51 += (afb[i][4] - M511) * (afb[i][4] - M511)
lamdabb[0]=S1/S11;lamdabb[1]=S2/S21;lamdabb[2]=S3/S31;lamdabb[3]=S4/S41;lamdabb[4]=S5/S51
lamdafb[0]=T1/T11;lamdafb[1]=T2/T21;lamdafb[2]=T3/T31;lamdafb[3]=T4/T41;lamdafb[4]=T5/T51
# girls factors
S1=0;S2=0;S3=0;S4=0;S5=0
S11=0;S21=0;S31=0;S41=0;S51=0
T1=0;T2=0;T3=0;T4=0;T5=0
T11=0;T21=0;T31=0;T41=0;T51=0
lamdabg=np.zeros(5)
lamdafg=np.zeros(5)

for i in range(50):
    lamdabg[0]=af[i][0]/abg[i][0]
    lamdabg[1] = af[i][1] / abg[i][1]
    lamdabg[2] = af[i][3] / abg[i][2]
    lamdabg[3] = af[i][5] / abg[i][3]
    lamdabg[4] = af[i][6] / abg[i][4]

    lamdafg[0] = af[i][0] / afg[i][0]
    lamdafg[1] = af[i][2] / afg[i][1]
    lamdafg[2] = af[i][4] / afg[i][2]
    lamdafg[3] = af[i][5] / afg[i][3]
    lamdafg[4] = af[i][7] / afg[i][4]

print(lamdabb)
print(lamdafb)
print(lamdabg)
print(lamdafg)